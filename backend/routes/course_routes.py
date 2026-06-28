from fastapi import APIRouter, Depends, Request
from motor.motor_asyncio import AsyncIOMotorDatabase

from models.course import CourseCreate
from models.lesson import LessonCreate
from models.resource import ResourceCreate
from services.course_service import CourseService
from services.auth_service import AuthService
from services.gamification_service import GamificationService
from config import get_settings

router = APIRouter(prefix="/api", tags=["courses"])


def get_course_service(request: Request) -> CourseService:
    return request.app.state.course_service


def get_auth_service(request: Request) -> AuthService:
    return request.app.state.auth_service


@router.post("/courses")
async def create_course(
    data: CourseCreate,
    request: Request,
    auth: AuthService = Depends(get_auth_service),
    course_svc: CourseService = Depends(get_course_service),
):
    user = await auth.require_role(request, "teacher")
    return await course_svc.create_course(
        data.title, data.description, data.subject, data.cover_color, user["id"], user["name"]
    )


@router.get("/courses")
async def list_courses(
    request: Request,
    auth: AuthService = Depends(get_auth_service),
    course_svc: CourseService = Depends(get_course_service),
    page: int = 1,
    limit: int = 12,
):
    user = await auth.get_current_user(request)
    skip = (max(page, 1) - 1) * limit
    items = await course_svc.list_courses(user["id"], skip=skip, limit=limit)
    total = await course_svc.count_courses()
    pages = max(1, (total + limit - 1) // limit)
    return {"items": items, "total": total, "page": page, "pages": pages}


@router.get("/courses/mine")
async def my_courses(
    request: Request,
    auth: AuthService = Depends(get_auth_service),
    course_svc: CourseService = Depends(get_course_service),
):
    user = await auth.get_current_user(request)
    return await course_svc.get_my_courses(user["id"], user["role"])


@router.get("/courses/{course_id}")
async def get_course(
    course_id: str,
    request: Request,
    auth: AuthService = Depends(get_auth_service),
    course_svc: CourseService = Depends(get_course_service),
):
    user = await auth.get_current_user(request)
    return await course_svc.get_course(course_id, user["id"])


@router.post("/courses/{course_id}/enroll")
async def enroll(
    course_id: str,
    request: Request,
    auth: AuthService = Depends(get_auth_service),
    course_svc: CourseService = Depends(get_course_service),
):
    user = await auth.require_role(request, "student")
    result = await course_svc.enroll_student(course_id, user["id"])
    if result.get("ok") and not result.get("already"):
        from services.notification_service import NotificationService
        notif_svc = NotificationService(request.app.state.db)
        course = await request.app.state.db.courses.find_one({"id": course_id}, {"_id": 0})
        if course:
            await notif_svc.create(
                user["id"],
                "Inscripción exitosa",
                f"Te inscribiste en {course['title']}",
                f"/courses/{course_id}",
            )
    return result


@router.delete("/courses/{course_id}")
async def delete_course(
    course_id: str,
    request: Request,
    auth: AuthService = Depends(get_auth_service),
    course_svc: CourseService = Depends(get_course_service),
):
    user = await auth.require_role(request, "teacher")
    return await course_svc.delete_course(course_id, user["id"])


@router.post("/courses/{course_id}/lessons")
async def create_lesson(
    course_id: str,
    data: LessonCreate,
    request: Request,
    auth: AuthService = Depends(get_auth_service),
):
    user = await auth.require_role(request, "teacher")
    db: AsyncIOMotorDatabase = request.app.state.db

    course = await db.courses.find_one({"id": course_id})
    if not course or course["teacher_id"] != user["id"]:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Not your course")

    import uuid
    from datetime import datetime, timezone
    doc = {
        "id": str(uuid.uuid4()),
        "course_id": course_id,
        "title": data.title,
        "content": data.content,
        "order": data.order,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.lessons.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.get("/courses/{course_id}/lessons")
async def list_lessons(
    course_id: str,
    request: Request,
    auth: AuthService = Depends(get_auth_service),
):
    await auth.get_current_user(request)
    db: AsyncIOMotorDatabase = request.app.state.db
    return await db.lessons.find({"course_id": course_id}, {"_id": 0}).sort("order", 1).to_list(200)


@router.delete("/lessons/{lesson_id}")
async def delete_lesson(
    lesson_id: str,
    request: Request,
    auth: AuthService = Depends(get_auth_service),
):
    user = await auth.require_role(request, "teacher")
    db: AsyncIOMotorDatabase = request.app.state.db

    lesson = await db.lessons.find_one({"id": lesson_id})
    if not lesson:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Not found")

    course = await db.courses.find_one({"id": lesson["course_id"]})
    if course["teacher_id"] != user["id"]:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Forbidden")

    await db.lessons.delete_one({"id": lesson_id})
    return {"ok": True}


@router.post("/courses/{course_id}/resources")
async def create_resource(
    course_id: str,
    data: ResourceCreate,
    request: Request,
    auth: AuthService = Depends(get_auth_service),
):
    user = await auth.require_role(request, "teacher")
    db: AsyncIOMotorDatabase = request.app.state.db

    course = await db.courses.find_one({"id": course_id})
    if not course or course["teacher_id"] != user["id"]:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Not your course")

    import uuid
    from datetime import datetime, timezone
    doc = {
        "id": str(uuid.uuid4()),
        "course_id": course_id,
        "title": data.title,
        "type": data.type,
        "url": data.url,
        "file_id": data.file_id,
        "description": data.description or "",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.resources.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.get("/courses/{course_id}/resources")
async def list_resources(
    course_id: str,
    request: Request,
    auth: AuthService = Depends(get_auth_service),
):
    await auth.get_current_user(request)
    db: AsyncIOMotorDatabase = request.app.state.db
    return await db.resources.find({"course_id": course_id}, {"_id": 0}).to_list(200)


@router.delete("/resources/{resource_id}")
async def delete_resource(
    resource_id: str,
    request: Request,
    auth: AuthService = Depends(get_auth_service),
):
    user = await auth.require_role(request, "teacher")
    db: AsyncIOMotorDatabase = request.app.state.db

    r = await db.resources.find_one({"id": resource_id})
    if not r:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Not found")

    course = await db.courses.find_one({"id": r["course_id"]})
    if course["teacher_id"] != user["id"]:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Forbidden")

    await db.resources.delete_one({"id": resource_id})
    return {"ok": True}


@router.get("/courses/{course_id}/students")
async def list_students(
    course_id: str,
    request: Request,
    auth: AuthService = Depends(get_auth_service),
):
    user = await auth.require_role(request, "teacher")
    db: AsyncIOMotorDatabase = request.app.state.db

    course = await db.courses.find_one({"id": course_id})
    if not course or course["teacher_id"] != user["id"]:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Not your course")

    enrollments = await db.enrollments.find({"course_id": course_id}).to_list(200)
    activities = await db.activities.find({"course_id": course_id}).to_list(200)

    students = []
    for e in enrollments:
        student = await db.users.find_one({"id": e["student_id"]}, {"_id": 0, "password": 0})
        if not student:
            continue

        submissions = await db.submissions.find({"student_id": e["student_id"]}).to_list(200)
        course_submissions = [s for s in submissions if any(a["id"] == s.get("activity_id") for a in activities)]

        submitted_activity_ids = {s.get("activity_id") for s in course_submissions}
        missing = [a for a in activities if a["id"] not in submitted_activity_ids]

        graded = [s for s in course_submissions if s.get("status") == "graded"]
        total_score = sum(s.get("score", 0) for s in graded)
        max_possible = sum(a.get("max_points", 100) for a in activities if a["id"] in {s.get("activity_id") for s in graded})

        students.append({
            "id": student["id"],
            "name": student.get("name", ""),
            "email": student.get("email", ""),
            "enrolled_at": e.get("enrolled_at", ""),
            "submissions_count": len(course_submissions),
            "graded_count": len(graded),
            "total_score": total_score,
            "max_possible": max_possible,
            "average": round(total_score / max_possible * 100, 1) if max_possible > 0 else 0,
            "missing_count": len(missing),
            "missing_activities": [{"id": a["id"], "title": a["title"], "type": a["type"]} for a in missing],
        })

    return students


@router.get("/courses/{course_id}/students/export")
async def export_students(
    course_id: str,
    request: Request,
    auth: AuthService = Depends(get_auth_service),
):
    user = await auth.require_role(request, "teacher")
    db: AsyncIOMotorDatabase = request.app.state.db

    course = await db.courses.find_one({"id": course_id})
    if not course or course["teacher_id"] != user["id"]:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Not your course")

    from fastapi.responses import StreamingResponse
    import openpyxl
    import io

    enrollments = await db.enrollments.find({"course_id": course_id}).to_list(200)
    activities = await db.activities.find({"course_id": course_id}).to_list(200)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    headers = ["Nombre", "Notas", "Promedio (%)", "Faltantes", "Actividades Faltantes"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="1F5A2A", end_color="1F5A2A", fill_type="solid")

    for e in enrollments:
        student = await db.users.find_one({"id": e["student_id"]}, {"_id": 0, "password": 0})
        if not student:
            continue

        submissions = await db.submissions.find({"student_id": e["student_id"]}).to_list(200)
        course_submissions = [s for s in submissions if any(a["id"] == s.get("activity_id") for a in activities)]

        submitted_activity_ids = {s.get("activity_id") for s in course_submissions}
        missing = [a for a in activities if a["id"] not in submitted_activity_ids]

        graded = [s for s in course_submissions if s.get("status") == "graded"]
        total_score = sum(s.get("score", 0) for s in graded)
        max_possible = sum(a.get("max_points", 100) for a in activities if a["id"] in {s.get("activity_id") for s in graded})
        average = round(total_score / max_possible * 100, 1) if max_possible > 0 else 0

        missing_titles = ", ".join([a["title"] for a in missing]) if missing else "Ninguna"

        ws.append([
            student.get("name", ""),
            f"{total_score}/{max_possible}",
            average,
            len(missing),
            missing_titles,
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"estudiantes_{course.get('title', 'curso')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
